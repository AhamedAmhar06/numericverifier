"""Enhanced file parser for CSV and Excel P&L tables.

Produces standard verifier table JSON:
{
  "type": "table",
  "content": {
    "caption": "...",
    "columns": ["", "FY2023", "FY2022"],
    "rows": [["Revenue", "383285", "394328"], ...]
  }
}
"""
import csv
import io
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

# Keywords used to identify P&L sheets in Excel workbooks
_PNL_KEYWORDS = (
    "income",
    "operations",
    "profit",
    "loss",
    "financial results",
    "p&l",
    "earnings",
    "statement",
)

# Pattern for scale notes like "in millions", "in thousands of $"
_SCALE_RE = re.compile(
    r"in\s+(thousands?|millions?|billions?)|"
    r"(thousands?|millions?|billions?)\s+of\s+[\$£€]?",
    re.IGNORECASE,
)

# Pattern for period-like column headers
_PERIOD_RE = re.compile(r"(FY|Q[1-4]|20\d{2}|19\d{2})", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_file(
    source: Union[str, "Path", bytes],
    filename: str = "",
    file_ext: str = "",
) -> Dict[str, Any]:
    """Parse a CSV or Excel file into standard verifier table JSON.

    Args:
        source:    File path, raw bytes, or text string.
        filename:  Original filename — used for caption and extension detection.
        file_ext:  Explicit extension (.csv / .xlsx / .xls) when not in filename.

    Returns:
        {"type": "table", "content": {"caption": str, "columns": [...], "rows": [...]}}
    """
    ext = file_ext.lower() if file_ext else Path(filename).suffix.lower() if filename else ""
    caption = _caption_from_filename(filename)

    if ext in (".xlsx", ".xls"):
        content = _parse_excel(source, caption)
    else:
        # Default / .csv
        content = _parse_csv(source, caption)

    return {"type": "table", "content": content}


# ---------------------------------------------------------------------------
# Caption helper
# ---------------------------------------------------------------------------

def _caption_from_filename(filename: str) -> str:
    if not filename:
        return ""
    stem = Path(filename).stem
    # CamelCase and underscore/hyphen separators → readable words
    readable = re.sub(r"([a-z])([A-Z])", r"\1 \2", stem)
    readable = re.sub(r"[_\-]+", " ", readable).strip()
    return readable


# ---------------------------------------------------------------------------
# Number cleanup helpers
# ---------------------------------------------------------------------------

def _clean_cell(s: str, is_label: bool = False) -> str:
    """Normalise a table cell value for the verifier."""
    if is_label:
        return s.strip()
    s = s.strip()
    if not s:
        return s
    # Parentheses negatives: (1,234) → -1234
    m = re.match(r"^\(([0-9,.\s]+)\)$", s)
    if m:
        inner = m.group(1).replace(",", "").strip()
        return f"-{inner}"
    # Remove thousands-separator commas from numbers: 1,234,567 → 1234567
    if re.match(r"^-?[\d,]+(\.\d+)?%?$", s):
        return s.replace(",", "")
    return s


def _looks_numeric(s: str) -> bool:
    """Return True if the cell value is likely a number."""
    s = re.sub(r"^\((.+)\)$", r"-\1", s.strip())
    s = s.replace(",", "").rstrip("%")
    try:
        float(s)
        return True
    except ValueError:
        return bool(re.match(r"^-?[\d.]+[KMBkmb]?$", s))


# ---------------------------------------------------------------------------
# CSV parser
# ---------------------------------------------------------------------------

def _parse_csv(
    source: Union[str, "Path", bytes],
    caption: str,
) -> Dict[str, Any]:
    """Parse a CSV source into {caption, columns, rows}."""
    if isinstance(source, bytes):
        text = source.decode("utf-8-sig", errors="replace")
        fh: Any = io.StringIO(text)
    elif isinstance(source, Path) or (isinstance(source, str) and "\n" not in source and Path(source).exists()):
        with open(str(source), newline="", encoding="utf-8-sig") as f:
            text = f.read()
        fh = io.StringIO(text)
    else:
        # Treat as raw CSV text
        fh = io.StringIO(str(source) if not isinstance(source, str) else source)

    reader = csv.reader(fh)
    rows_raw: List[List[str]] = [
        row for row in reader if any(c.strip() for c in row)
    ]

    if not rows_raw:
        raise ValueError("CSV is empty or has no data rows")

    # Auto-detect orientation
    # Layout A (normal):  row-0 = period headers, col-0 = line item labels
    # Layout B (transposed): col-0 = period labels, remaining cols = line items
    first_cell = rows_raw[0][0].strip() if rows_raw[0] else ""
    is_layout_b = bool(_PERIOD_RE.match(first_cell)) and len(rows_raw) >= 2

    if is_layout_b:
        columns, rows = _transpose_csv(rows_raw)
    else:
        columns = [c.strip() for c in rows_raw[0]]
        rows = _build_rows(rows_raw[1:], len(columns))

    return {"caption": caption, "columns": columns, "rows": rows}


def _build_rows(raw_data: List[List[str]], n_cols: int) -> List[List[str]]:
    rows = []
    for row in raw_data:
        padded = list(row) + [""] * max(0, n_cols - len(row))
        cleaned = [
            _clean_cell(c, is_label=(i == 0))
            for i, c in enumerate(padded[:n_cols])
        ]
        rows.append(cleaned)
    return rows


def _transpose_csv(raw_rows: List[List[str]]):
    """Transpose Layout-B (periods as rows) into Layout-A (periods as columns)."""
    max_cols = max(len(r) for r in raw_rows)
    padded = [r + [""] * (max_cols - len(r)) for r in raw_rows]
    transposed = [list(col) for col in zip(*padded)]
    columns = [c.strip() for c in transposed[0]]
    rows = _build_rows(transposed[1:], len(columns))
    return columns, rows


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------

def _parse_excel(
    source: Union[str, "Path", bytes],
    caption: str,
) -> Dict[str, Any]:
    """Parse an Excel workbook into {caption, columns, rows, scale_note?}."""
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel parsing. "
            "Install with: pip install openpyxl"
        )

    if isinstance(source, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(Path(source), read_only=True, data_only=True)

    ws = _find_pnl_sheet(wb)
    scale_note = _extract_scale_note(ws)

    # Collect non-empty rows
    all_rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [_cell_to_str(c) for c in row]
        if any(c for c in cells):
            all_rows.append(cells)
    wb.close()

    if not all_rows:
        raise ValueError("Excel P&L sheet is empty")

    # Find the header row (first row where ≥1 col matches a period pattern)
    header_idx = _find_header_row(all_rows)
    header_row = all_rows[header_idx]
    columns = [c.strip() for c in header_row]
    rows = _build_rows(all_rows[header_idx + 1:], len(columns))

    content: Dict[str, Any] = {"caption": caption, "columns": columns, "rows": rows}
    if scale_note:
        content["scale_note"] = scale_note
    return content


def _find_pnl_sheet(wb: Any) -> Any:
    """Return the most likely P&L sheet by keyword-matching sheet names."""
    for name in wb.sheetnames:
        lower = name.lower()
        if any(kw in lower for kw in _PNL_KEYWORDS):
            return wb[name]
    active = wb.active
    return active if active is not None else wb[wb.sheetnames[0]]


def _extract_scale_note(ws: Any) -> Optional[str]:
    """Scan the first few rows for a scale note like 'in millions'."""
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=8)):
        for cell in row:
            if cell is None:
                continue
            s = str(cell)
            if _SCALE_RE.search(s):
                return s.strip()
        if i >= 7:
            break
    return None


def _find_header_row(rows: List[List[str]]) -> int:
    """Return the index of the row that contains period column headers."""
    for i, row in enumerate(rows[:12]):
        period_count = sum(1 for c in row[1:] if _PERIOD_RE.search(str(c)))
        if period_count >= 1:
            return i
    return 0


def _cell_to_str(cell: Any) -> str:
    if cell is None:
        return ""
    return str(cell).strip()
