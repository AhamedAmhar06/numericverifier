"""Parse an XLSX file into verifier-compatible table JSON."""
from pathlib import Path
from typing import Dict, Any, Union


def parse_excel_pnl(source: Union[str, Path], sheet_name: str = None) -> Dict[str, Any]:
    """Read the first (or named) sheet of an XLSX workbook and return ``{columns, rows, units}``."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX parsing. "
            "Install it with: pip install openpyxl>=3.1.0"
        )

    path = Path(source)
    if not path.exists():
        raise FileNotFoundError(f"XLSX file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws is None:
        wb.close()
        raise ValueError("Workbook has no active sheet")

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([_cell_to_str(c) for c in row])
    wb.close()

    if not all_rows:
        raise ValueError("XLSX sheet is empty")

    columns = all_rows[0]
    n_cols = len(columns)
    rows = []
    for row in all_rows[1:]:
        padded = list(row) + [""] * max(0, n_cols - len(row))
        rows.append(padded[:n_cols])

    return {"columns": columns, "rows": rows, "units": {}}


def _cell_to_str(cell) -> str:
    if cell is None:
        return ""
    return str(cell).strip()
